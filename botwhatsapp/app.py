import openpyxl
from openpyxl import load_workbook
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import os

webbrowser.open('https://web.whatsapp.com/')
sleep(30)

#ler planilha e guardar informações sobre nome, telefone
workbook = load_workbook('lista-rede-01.xlsx')
print(workbook.sheetnames)
pagina_clientes = workbook['Planilha2']

for linha in pagina_clientes.iter_rows(min_row=1):
    #nome e telefone
    telefone = linha[1].value
    nome = linha[2].value

    mensagem = f'Olá {nome} você é um amigo, amigo. Este é um teste de automação'

    try:
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
        webbrowser.open(link_mensagem_whatsapp)
        sleep(10)
        pyautogui.press('enter')
        sleep(3)
        #seta = pyautogui.locateCenterOnScreen('seta.png')
        #sleep(2)
        #pyautogui.click(seta[0],seta[1])
        #sleep(2)
        pyautogui.hotkey('ctrl','w')
        sleep(2)
    except:
        print(f'Não foi possível enviar mensagem para {nome}')
        with open('erros.csv','a',newline='',encoding='utf-8') as arquivo:
            arquivo.write(f'{nome},{telefone}{os.linesep}')

